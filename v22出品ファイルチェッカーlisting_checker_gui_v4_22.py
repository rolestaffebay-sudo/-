# listing_checker_gui.py
# ------------------------------------------------------------
# 初回のみ:
#   py -m pip install openpyxl pandas requests playwright
#   py -m playwright install chromium
#
# 起動:
#   py listing_checker_gui.py
# ------------------------------------------------------------

# -*- coding: utf-8 -*-

import os
import re
import csv
import shutil
import threading
import datetime as dt
import webbrowser
import random
import json

SETTINGS_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "listing_checker_settings.json"
)

from dataclasses import dataclass, asdict
from typing import List, Optional, Tuple, Dict, Any
from queue import Queue, Empty

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

def load_settings() -> dict:
    if not os.path.exists(SETTINGS_FILE):
        return {}
    try:
        with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def save_settings(data: dict) -> None:
    try:
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _startup_die(msg: str) -> None:
    """起動直後の依存関係エラー等を、GUIでも分かる形で表示して終了する。"""
    try:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("起動エラー", msg)
        try:
            root.destroy()
        except Exception:
            pass
    except Exception:
        # Tkが使えない場合は標準出力へ
        print(msg)
    raise SystemExit(1)

try:
    import openpyxl
except ModuleNotFoundError:
    _startup_die("openpyxl が見つかりません。\n\n以下を実行してください:\n  py -m pip install openpyxl")

try:
    import pandas as pd
except ModuleNotFoundError:
    _startup_die("pandas が見つかりません。\n\n以下を実行してください:\n  py -m pip install pandas")

try:
    import requests
except ModuleNotFoundError:
    _startup_die("requests が見つかりません。\n\n以下を実行してください:\n  py -m pip install requests")

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError
except ModuleNotFoundError:
    _startup_die("playwright が見つかりません。\n\n以下を実行してください:\n  py -m pip install playwright\n  py -m playwright install chromium")


# -----------------------------
# Template columns (1-based)
# -----------------------------
COL_SKU = 1          # A
COL_PRICE = 2        # B  (foreign price)
COL_QTY = 3          # C
COL_ASIN = 4         # D

ASIN_RE = re.compile(r"^[A-Z0-9]{10}$")

CUSTOMS_FIXED_YEN_DEFAULT = 150.0

COUNTRY_MULTIPLIER = {
    "US": 1.692,
    "CA": 1.56,
    "MX": 2.463
}

FX_REFRESH_MS = 30 * 60 * 1000

def _default_pw_profile_dir() -> str:
    """
    Playwright の永続プロファイル保存先（user_data_dir）。
    実行時のカレントディレクトリに依存すると System32 等に作ろうとして失敗するため、
    まずは「この .py と同じフォルダ/pw_profile」を試し、権限NGなら LocalAppData に退避します。
    """
    here = os.path.dirname(os.path.abspath(__file__))
    preferred = os.path.join(here, "pw_profile")
    try:
        os.makedirs(preferred, exist_ok=True)
        # 書き込み確認（権限/保護フォルダ対策）
        test_path = os.path.join(preferred, ".write_test")
        with open(test_path, "w", encoding="utf-8") as f:
            f.write("ok")
        os.remove(test_path)
        return preferred
    except Exception:
        pass

    base = os.environ.get("LOCALAPPDATA") or os.path.expanduser("~")
    alt = os.path.join(base, "listing_checker", "pw_profile")
    os.makedirs(alt, exist_ok=True)
    return alt

PW_PROFILE_DIR = _default_pw_profile_dir()
# ---- Timeouts ----
AMZ_NAV_TIMEOUT_MS = 30000      # ページ遷移30秒（初回起動・低速環境対策）
AMZ_HARD_TIMEOUT_MS = 60000     # 全体60秒でGUI解除（初回起動・低速環境対策）

# 高速化：重いリソースのみブロック（stylesheetはブロックしない）
AMZ_BLOCK_RESOURCE_TYPES = {"image", "font", "media"}

DEFAULT_HEADLESS = True


def clean_text(v) -> str:
    if v is None:
        return ""
    return str(v).replace("\r", "").replace("\n", "").strip()


def try_float(v: str) -> Optional[float]:
    v = clean_text(v)
    if v == "":
        return None
    v = v.replace(",", "")
    try:
        return float(v)
    except ValueError:
        return None


def find_header_row(ws, search_rows: int = 80) -> Optional[int]:
    for r in range(1, min(search_rows, ws.max_row) + 1):
        row_values = [clean_text(ws.cell(r, c).value).lower() for c in range(1, min(ws.max_column, 40) + 1)]
        if "sku" in row_values and "product-id" in row_values:
            return r
    return None


def build_data_row_list(ws, header_row: int) -> List[int]:
    data_rows: List[int] = []
    start = header_row + 1
    for r in range(start, ws.max_row + 1):
        asin = clean_text(ws.cell(r, COL_ASIN).value)
        sku = clean_text(ws.cell(r, COL_SKU).value)
        price = clean_text(ws.cell(r, COL_PRICE).value)

        if asin == "" and sku == "" and price == "":
            continue
        if asin.lower() == "product-id" or sku.lower() == "sku":
            continue
        if asin != "":
            data_rows.append(r)
    return data_rows


def parse_sku_cost_yen(sku: str) -> Optional[float]:
    sku = clean_text(sku)
    if sku == "":
        return None
    parts = sku.split("-")
    if len(parts) < 2:
        return None
    return try_float(parts[1])


def parse_sku_shipping_yen(sku: str, country: str) -> Tuple[float, float]:
    """
    SKUから「通常送料（円）」と「MX小型包装物の値（円）」を返す。

    通常送料（円）:
      - housou / biggg / kogata いずれも parts[2] × 100 を送料（円）として扱う（従来通り）

    MX小型包装物（MXのみ、housou / kogata のみ適用）:
      - SKUの parts[3] を「小型包装物の値（円）」として返す（例：...-28-... -> 28）
      - この値自体は base_yen には加算しない（小型包装物送料は別式で外貨側に加算するため）
      - US/CA では常に 0 を返す
    """
    sku = clean_text(sku)
    if sku == "":
        return 0.0, 0.0
    parts = sku.split("-")
    if len(parts) < 3:
        return 0.0, 0.0

    sku_type = parts[0].lower()
    ship_factor = try_float(parts[2])
    ship_yen = (ship_factor * 100.0) if ship_factor is not None else 0.0

    mx_small_yen = 0.0
    if country.upper() == "MX" and sku_type in ("housou", "kogata") and len(parts) >= 4:
        # MX小型包装物の「値」（円）。例：...-28-... -> 28
        small_val = try_float(parts[3])
        mx_small_yen = float(small_val) if small_val is not None else 0.0

    return ship_yen, mx_small_yen


def fetch_fx_rates_jpy_per_unit(timeout_sec: int = 10) -> Optional[Dict[str, float]]:
    try:
        url = "https://open.er-api.com/v6/latest/JPY"
        r = requests.get(url, timeout=timeout_sec)
        r.raise_for_status()
        data = r.json()
        rates = data.get("rates", {})

        usd_per_jpy = rates.get("USD")
        cad_per_jpy = rates.get("CAD")
        mxn_per_jpy = rates.get("MXN")
        if not usd_per_jpy or not cad_per_jpy or not mxn_per_jpy:
            return None

        usd_jpy = 1.0 / float(usd_per_jpy)
        cad_jpy = 1.0 / float(cad_per_jpy)
        mxn_jpy = 1.0 / float(mxn_per_jpy)

        return {"USDJPY": usd_jpy, "CADJPY": cad_jpy, "MXNJPY": mxn_jpy}
    except Exception:
        return None


class AmazonAutoFetcher:
    def __init__(self, profile_dir: str, headless: bool):
        self.profile_dir = profile_dir
        self.headless = headless
        self._pw = None
        self._context = None
        self._page = None
        self._route_set = False

    def start(self):
        if self._page is not None:
            return

        ua = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
              "AppleWebKit/537.36 (KHTML, like Gecko) "
              "Chrome/122.0.0.0 Safari/537.36")

        self._pw = sync_playwright().start()
        self._context = self._pw.chromium.launch_persistent_context(
            user_data_dir=self.profile_dir,
            headless=self.headless,
            locale="ja-JP",
            timezone_id="Asia/Tokyo",
            user_agent=ua,
            args=["--disable-blink-features=AutomationControlled"],
        )
        self._page = self._context.new_page()
        self._page.set_default_navigation_timeout(AMZ_NAV_TIMEOUT_MS)
        self._page.set_default_timeout(AMZ_NAV_TIMEOUT_MS)

        if not self._route_set:
            def _route_handler(route):
                try:
                    if route.request.resource_type in AMZ_BLOCK_RESOURCE_TYPES:
                        route.abort()
                    else:
                        route.continue_()
                except Exception:
                    try:
                        route.continue_()
                    except Exception:
                        pass

            try:
                self._page.route("**/*", _route_handler)
                self._route_set = True
            except Exception:
                self._route_set = False

    def stop(self):
        try:
            if self._context:
                self._context.close()
        finally:
            self._context = None
            self._page = None
            if self._pw:
                self._pw.stop()
                self._pw = None
    @staticmethod
    def _first_yen_in_text(text: str) -> Optional[int]:
        """テキストから「円価格」を抽出する。

        誤検出（例: "1個" の 1）を避けるため、以下の優先順位で抽出する:
          1) ￥ / ¥ を含む価格
          2) "円" を含む価格
          3) 通貨記号が無い場合は「3桁以上」かつ「100以上」の数値のみ採用（1や2などを排除）
        """
        if not text:
            return None

        # 1) 明示的な通貨記号
        m = re.search(r"[￥¥]\s*([\d,]+)", text)
        if m:
            v = int(m.group(1).replace(",", ""))
            return v if 0 <= v <= 9_999_999 else None

        # 2) 「円」
        m = re.search(r"([\d,]+)\s*円", text)
        if m:
            v = int(m.group(1).replace(",", ""))
            return v if 0 <= v <= 9_999_999 else None

        # 3) 最終フォールバック（過剰に拾わない）
        #    例: "4,519" はOK / "989" もOK / "1" はNG
        m2 = re.search(r"\b(\d{1,3}(?:,\d{3})+|\d{3,})\b", text)
        if m2:
            v = int(m2.group(1).replace(",", ""))
            if v >= 100 and v <= 9_999_999:
                return v

        return None

    @staticmethod
    def _parse_shipping_yen_from_text(text: str) -> Optional[int]:
        if not text:
            return None
        if "送料無料" in text or "無料配送" in text:
            return 0
        patterns = [
            r"(配送料|送料)\s*[:：]?\s*\+?\s*[￥¥]\s*([\d,]+)",
            r"\+\s*[￥¥]\s*([\d,]+)\s*(配送料|送料)",
            r"[￥¥]\s*([\d,]+)\s*(配送料|送料)",
        ]
        for p in patterns:
            m = re.search(p, text)
            if not m:
                continue
            nums = re.findall(r"[\d,]+", m.group(0))
            if nums:
                return int(nums[0].replace(",", ""))
        return None

    @staticmethod
    def _shipped_by_amazon_from_text(text: str) -> bool:
        if not text:
            return False
        t = str(text)
        keys = ["出荷元", "Ships from", "Dispatches from"]
        if any(k in t for k in keys) and ("Amazon" in t or "Amazon.co.jp" in t or "amazon" in t.lower()):
            return True
        return False

    @staticmethod
    def _delivery_days_from_text(text: str, today: Optional[dt.date] = None) -> Optional[int]:
        """配送目安テキストから「最長側」の配送日数（整数）を推定する。

        - 日付表現が複数出る場合（例: 1月10日-1月12日）は遅い方（最大delta）を採用。
        - 形式:
            * 2026年1月8日
            * 1月8日
            * 1/8, 01/08, 2026/1/8
            * 今日 / 明日
            * 1-2日, 1～2日, 2日
        """
        if not text:
            return None
        if today is None:
            today = dt.date.today()

        # 今日 / 明日
        if "明日" in text:
            return 1
        if "今日" in text:
            return 0

        candidates: List[dt.date] = []

        # yyyy年m月d日
        for m in re.finditer(r"(\d{4})年(\d{1,2})月(\d{1,2})日", text):
            try:
                y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
                candidates.append(dt.date(y, mo, d))
            except Exception:
                pass

        # m月d日（年省略）
        for m in re.finditer(r"(\d{1,2})月(\d{1,2})日", text):
            try:
                mo, d = int(m.group(1)), int(m.group(2))
                y = today.year
                # 年跨ぎの可能性を雑に補正（年末〜年始）
                if mo < today.month - 6:
                    y += 1
                candidates.append(dt.date(y, mo, d))
            except Exception:
                pass

        # yyyy/m/d or m/d
        for m in re.finditer(r"(?:(\d{4})\s*/\s*)?(\d{1,2})\s*/\s*(\d{1,2})", text):
            try:
                y = int(m.group(1)) if m.group(1) else today.year
                mo, d = int(m.group(2)), int(m.group(3))
                if (not m.group(1)) and mo < today.month - 6:
                    y += 1
                candidates.append(dt.date(y, mo, d))
            except Exception:
                pass

        if candidates:
            deltas = [(c - today).days for c in candidates]
            deltas = [d for d in deltas if d >= 0]
            if deltas:
                mx = max(deltas)
                if 0 <= mx <= 60:
                    return mx

        # 「1～2日」「1-2日」など
        m2 = re.search(r"(\d+)\s*[～\-]\s*(\d+)\s*日", text)
        if m2:
            hi = int(m2.group(2))
            return hi if 0 <= hi <= 30 else None

        # 「2日」など
        m3 = re.search(r"(\d+)\s*日", text)
        if m3:
            n = int(m3.group(1))
            return n if 0 <= n <= 30 else None

        return None

    def _dismiss_cookie_consent(self) -> bool:
        if self._page is None:
            return False
        candidates = [
            "同意する", "同意して続行", "すべてのCookieを受け入れる",
            "Accept", "Accept Cookies", "Allow all cookies"
        ]
        for name in candidates:
            try:
                self._page.get_by_role("button", name=re.compile(name)).click(timeout=1200)
                return True
            except Exception:
                pass
        return False

    def _extract_selected_tile_price_yen(self) -> Optional[int]:
        """
        画像のような「選択中（青枠）」タイル価格を優先取得。
        """
        if self._page is None:
            return None

        selectors = [
            "#twister_feature_div .swatchElement.selected",
            "#twister_feature_div [aria-checked='true']",
            "#twister_feature_div [aria-selected='true']",
            "#twister_feature_div [aria-pressed='true']",
            "[id*='variation'] .swatchElement.selected",
            "[id*='variation'] [aria-checked='true']",
            "[id*='variation'] [aria-selected='true']",
            "[id*='variation'] [aria-pressed='true']",
            ".a-button-toggle.a-button-selected",
            ".a-button-selected",
            ".a-box.a-box-selected",
            "[aria-current='true']",
        ]

        for sel in selectors:
            try:
                el = self._page.query_selector(sel)
                if not el:
                    continue
                txt = clean_text(el.inner_text())
                yen = self._first_yen_in_text(txt)
                if yen is not None:
                    return yen
                try:
                    full = clean_text(el.evaluate("e => e.innerText"))
                    yen2 = self._first_yen_in_text(full)
                    if yen2 is not None:
                        return yen2
                except Exception:
                    pass
            except Exception:
                continue

        return None

    
    def _fetch_best_offer_from_offer_listing(self, asin: str, sku_cost_yen: Optional[float]) -> Optional[Tuple[int, int, Optional[int], bool, bool, str]]:
        """おすすめの出品（BuyBox）が無い／取得できない場合に、出品一覧ページから候補を選ぶ。

        選定ルール（SKU情報に近い出品を優先）:
          - sku_cost_yen が取れる場合: |Amazon商品価格 - sku_cost_yen| が最小の出品を優先
          - 取れない場合: Amazon(商品+配送料) 合計が最小の出品を優先
          - 同点なら: 出荷元Amazonを優先 → 合計が安い → 配送日数が短い
        """
        if self._page is None:
            return None

        offer_url = f"https://www.amazon.co.jp/gp/offer-listing/{asin}?condition=new&th=1&psc=1"
        debug_parts: List[str] = []
        try:
            self._page.goto(offer_url, wait_until="domcontentloaded")
        except Exception as e:
            return None

        # Cookie同意（必要なら）
        try:
            if self._dismiss_cookie_consent():
                debug_parts.append("cookie_clicked")
                self._page.wait_for_timeout(400)
        except Exception:
            pass

        # 代表的なオファーDOM（複数パターンに対応）
        offer_selectors = [
            "div.olpOffer",
            "div#aod-offer-list div.aod-offer",
            "div.aod-offer",
            "div[role='listitem']",
        ]
        offer_els = []
        used_sel = ""
        for sel in offer_selectors:
            try:
                els = self._page.query_selector_all(sel)
                if els:
                    offer_els = els
                    used_sel = sel
                    break
            except Exception:
                continue

        if not offer_els:
            return None

        debug_parts.append(f"offers:{used_sel}:{min(len(offer_els), 10)}")

        best = None  # (score_tuple, item, ship, days, shipped_by_amazon, is_used, raw_snip)
        for el in offer_els[:10]:
            try:
                raw = clean_text(el.inner_text())
            except Exception:
                continue
            if not raw:
                continue

            # 価格
            price = None
            try:
                pel = el.query_selector("span.a-price span.a-offscreen")
                if pel:
                    price = self._first_yen_in_text(clean_text(pel.inner_text()))
            except Exception:
                pass
            if price is None:
                price = self._first_yen_in_text(raw)

            # 配送料（このプロジェクトのルール上、取れない場合は後でDELETEになるため候補から外す）
            ship = self._parse_shipping_yen_from_text(raw)

            # 配送日数
            days = self._delivery_days_from_text(raw)

            shipped_by_amazon = ("出荷元 Amazon" in raw) or ("Ships from Amazon" in raw) or ("Amazonが発送" in raw) or ("Amazon.co.jp が発送" in raw)
            is_used = ("中古" in raw) or ("中古品" in raw)

            if price is None or ship is None:
                # 価格/送料が取れない候補は除外（最終的に全候補が除外される場合は呼び出し側でDELETE扱い）
                continue

            total = price + ship
            if sku_cost_yen is not None:
                primary = abs(float(price) - float(sku_cost_yen))
            else:
                primary = float(total)

            days_score = days if days is not None else 999

            score = (primary, 0 if shipped_by_amazon else 1, total, days_score)
            snip = raw[:120]
            if best is None or score < best[0]:
                best = (score, price, ship, days, shipped_by_amazon, is_used, snip)

        if best is None:
            return None

        _, price, ship, days, shipped_by_amazon, is_used, snip = best
        dbg = "offer_selected" + ("|" + "|".join(debug_parts) if debug_parts else "")
        dbg += f" :: {snip}"
        return price, ship, days, shipped_by_amazon, is_used, dbg

    def fetch_item_ship_delivery(self, asin: str, sku_cost_yen: Optional[float] = None) -> Tuple[Optional[int], Optional[int], Optional[int], Optional[bool], Optional[bool], bool, str]:
        self.start()
        assert self._page is not None

        url = f"https://www.amazon.co.jp/dp/{asin}?th=1&psc=1"
        debug_parts: List[str] = []

        try:
            self._page.goto(url, wait_until="domcontentloaded")
        except PWTimeoutError:
            debug_parts.append("goto_timeout")
        except Exception as e:
            debug_parts.append(f"goto_err:{e}")

        if self._dismiss_cookie_consent():
            debug_parts.append("cookie_clicked")
            try:
                self._page.wait_for_timeout(600)
            except Exception:
                pass

        # 価格（選択タイル優先）
        item_price = self._extract_selected_tile_price_yen()
        if item_price is not None:
            debug_parts.append("price:selected_tile")
        else:
            price_selectors = [
                "span.a-price span.a-offscreen",
                "#corePriceDisplay_desktop_feature_div span.a-offscreen",
                "#priceblock_ourprice",
                "#priceblock_dealprice",
            ]
            for sel in price_selectors:
                try:
                    el = self._page.query_selector(sel)
                    if not el:
                        continue
                    txt = clean_text(el.inner_text())
                    py = self._first_yen_in_text(txt)
                    if py is not None:
                        item_price = py
                        debug_parts.append(f"price:{sel}")
                        break
                except Exception:
                    pass

        # 配送/配送料候補
        texts: List[str] = []
        selectors = [
            "#mir-layout-DELIVERY_BLOCK-slot-PRIMARY_DELIVERY_MESSAGE_LARGE span",
            "#mir-layout-DELIVERY_BLOCK-slot-SECONDARY_DELIVERY_MESSAGE_LARGE span",
            "#deliveryBlockMessage span",
            "#mir-layout-DELIVERY_BLOCK-slot-DELIVERY_MESSAGE span",
            "#mir-layout-DELIVERY_BLOCK-slot-DELIVERY_MESSAGE",
            "#mir-layout-DELIVERY_BLOCK-slot-PRIMARY_DELIVERY_MESSAGE_LARGE",
            "#mir-layout-DELIVERY_BLOCK-slot-SECONDARY_DELIVERY_MESSAGE_LARGE",
            "#deliveryBlockMessage",
            "#tabular-buybox",
            "#tabular-buybox tr",
            "#merchant-info",
            "#merchant-info span",
        ]
        for sel in selectors:
            try:
                els = self._page.query_selector_all(sel)
                for el in els[:10]:
                    t = clean_text(el.inner_text())
                    if t:
                        texts.append(t)
            except Exception:
                pass

        raw_text = " / ".join(texts[:6]) if texts else ""

        # 出荷元がAmazonかどうか（配送判定の免除用）
        shipped_by_amazon = False
        for t in texts:
            if self._shipped_by_amazon_from_text(t):
                shipped_by_amazon = True
                debug_parts.append("shipper:amazon")
                break
        if not shipped_by_amazon:
            try:
                html = self._page.content()
                if re.search(r"出荷元.{0,80}Amazon", html, re.DOTALL) or re.search(r"Ships from.{0,80}Amazon", html, re.DOTALL):
                    shipped_by_amazon = True
                    debug_parts.append("shipper:amazon_html")
            except Exception:
                pass


        # 商品状態（新品/中古）の判定：中古の場合は後段でDELETEにするためフラグ化
        is_used = None  # True=中古, False=新品, None=不明
        try:
            cond_texts: List[str] = []
            cond_selectors = [
                "#conditionText",
                "#conditionText span",
                "#tabular-buybox tr",
                "#tabular-buybox",
                "#buybox_feature_div",
                "#buyBoxInner",
                "#buybox",
            ]
            for sel in cond_selectors:
                try:
                    els = self._page.query_selector_all(sel)
                    for el in els[:10]:
                        t = clean_text(el.inner_text())
                        if t:
                            cond_texts.append(t)
                except Exception:
                    pass

            # (1) "状態" 行（tabular-buybox等）を最優先
            for t in cond_texts:
                if ("状態" in t) or ("コンディション" in t) or ("Condition" in t):
                    if re.search(r"(状態|コンディション|Condition)\s*[:：]?\s*新品", t):
                        is_used = False
                        debug_parts.append("cond:new")
                        break
                    if re.search(r"(状態|コンディション|Condition)\s*[:：]?\s*中古", t):
                        is_used = True
                        debug_parts.append("cond:used")
                        break
                    # 状態行に新品/中古が混在するケースは新品優先
                    if ("新品" in t) and ("中古" not in t):
                        is_used = False
                        debug_parts.append("cond:new")
                        break
                    if ("中古" in t) and ("新品" not in t):
                        is_used = True
                        debug_parts.append("cond:used")
                        break

            # (2) conditionText フォールバック（短文で新品/中古が出るケース）
            if is_used is None:
                for t in cond_texts[:2]:
                    if ("新品" in t) and ("中古" not in t):
                        is_used = False
                        debug_parts.append("cond:new_ct")
                        break
                    if ("中古" in t) and ("新品" not in t):
                        is_used = True
                        debug_parts.append("cond:used_ct")
                        break
        except Exception:
            pass

        delivery_days = None
        for t in texts:
            delivery_days = self._delivery_days_from_text(t)
            if delivery_days is not None:
                break
        if delivery_days is None:
            # 取得直後は配送ブロックがまだ更新されないケースがあるため、軽く待って再抽出
            try:
                self._page.wait_for_timeout(900)
                try:
                    self._page.evaluate("() => window.scrollBy(0, 600)")
                except Exception:
                    pass
            except Exception:
                pass

            texts_retry: List[str] = []
            for sel in selectors:
                try:
                    els = self._page.query_selector_all(sel)
                    for el in els[:10]:
                        t = clean_text(el.inner_text())
                        if t:
                            texts_retry.append(t)
                except Exception:
                    pass

            for t in texts_retry:
                delivery_days = self._delivery_days_from_text(t)
                if delivery_days is not None:
                    debug_parts.append("days:retry")
                    break

            # デバッグ表示用に、再抽出も含める
            if texts_retry:
                texts = texts + texts_retry

        page_ship = None
        for t in texts:
            page_ship = self._parse_shipping_yen_from_text(t)
            if page_ship is not None:
                debug_parts.append("ship:texts")
                break

        if page_ship is None:
            try:
                html = self._page.content()
                if "送料無料" in html or "無料配送" in html:
                    page_ship = 0
                    debug_parts.append("ship:html_free")
                else:
                    m = re.search(r"(配送料|送料)[^￥¥]{0,40}[￥¥]\s*([\d,]+)", html)
                    if m:
                        page_ship = int(m.group(2).replace(",", ""))
                        debug_parts.append("ship:html_regex")
            except Exception:
                pass

        
        # おすすめの出品（BuyBox）が無い場合はフラグ化（要望：この場合はDELETE）
        no_featured_offer = False
        try:
            html2 = self._page.content()
            # BuyBoxがある場合は通常ここに「カートに入れる」等が出る
            has_cart = False
            try:
                has_cart = (self._page.locator('#add-to-cart-button, input#add-to-cart-button, #buy-now-button').count() > 0)
            except Exception:
                has_cart = False
            if ('おすすめの出品はありません' in html2) or ('おすすめの出品がありません' in html2):
                no_featured_offer = True
            elif (('すべての購入オプション' in html2) or ('See All Buying Options' in html2)) and (not has_cart):
                # 購入オプション誘導が出ていて、かつカートボタンが無ければBuyBox無しとみなす
                no_featured_offer = True
        except Exception:
            pass
        
        if no_featured_offer:
            debug_parts.append('no_featured_offer')
            # 旧仕様：出品一覧(offer-listing)から代替取得
            # 現仕様：おすすめの出品が無い場合は DELETE とするため、代替取得は行わない
        
        debug = f"{'|'.join(debug_parts)} :: {raw_text[:140]}"
        return item_price, page_ship, delivery_days, shipped_by_amazon, is_used, no_featured_offer, debug


class AmazonFetchWorker(threading.Thread):
    """
    Playwright操作はこのスレッド1本のみで行う。
    in_q: 指示
    out_q: 結果
    """
    def __init__(self, out_q: "Queue[Dict[str, Any]]"):
        super().__init__(daemon=True)
        self.in_q: "Queue[Dict[str, Any]]" = Queue()
        self.out_q = out_q
        self._stop = threading.Event()
        self._headless = DEFAULT_HEADLESS
        self.fetcher: Optional[AmazonAutoFetcher] = None

    def set_headless(self, headless: bool):
        self.in_q.put({"type": "SET_HEADLESS", "headless": headless})

    def fetch(self, asin: str, token: int, sku_cost_yen: Optional[float] = None):
        self.in_q.put({"type": "FETCH", "asin": asin, "token": token, "sku_cost_yen": sku_cost_yen})

    def stop(self):
        self.in_q.put({"type": "STOP"})

    def _reset_fetcher(self):
        if self.fetcher is not None:
            try:
                self.fetcher.stop()
            except Exception:
                pass
        self.fetcher = None

    def run(self):
        while not self._stop.is_set():
            try:
                task = self.in_q.get(timeout=0.2)
            except Empty:
                continue

            ttype = task.get("type")

            if ttype == "STOP":
                self._stop.set()
                break

            if ttype == "SET_HEADLESS":
                self._headless = bool(task.get("headless"))
                self._reset_fetcher()
                continue

            if ttype == "FETCH":
                asin = task["asin"]
                sku_cost_yen = task.get("sku_cost_yen")
                token = task["token"]
                try:
                    if self.fetcher is None:
                        self.fetcher = AmazonAutoFetcher(profile_dir=PW_PROFILE_DIR, headless=self._headless)
                    item, ship, days, shipped_by_amazon, is_used, no_featured_offer, debug = self.fetcher.fetch_item_ship_delivery(asin, sku_cost_yen=sku_cost_yen)
                    self.out_q.put({
                        "token": token,
                        "asin": asin,
                        "ok": True,
                        "item": item,
                        "ship": ship,
                        "days": days,
                        "shipped_by_amazon": shipped_by_amazon,
                        "is_used": is_used,
                        "no_featured_offer": no_featured_offer,
                        "debug": debug
                    })
                except Exception as e:
                    self.out_q.put({
                        "token": token,
                        "asin": asin,
                        "ok": False,
                        "item": None,
                        "ship": None,
                        "days": None,
                        "shipped_by_amazon": None,
                        "no_featured_offer": False,
                        "debug": f"worker_err:{e}"
                    })
                continue

        self._reset_fetcher()


@dataclass
class CheckRecord:
    timestamp: str
    country: str
    excel_row: int
    asin: str
    sku: str

    listing_price_foreign: Optional[float]

    amazon_item_yen: Optional[float]
    amazon_page_ship_yen: Optional[float]
    amazon_total_yen: Optional[float]
    shipped_by_amazon: Optional[bool]
    is_used: Optional[bool]
    no_featured_offer: Optional[bool]

    sku_ship_yen: Optional[float]
    mx_small_yen: Optional[float]
    mx_small_ship_foreign: Optional[float]
    customs_fixed_yen: Optional[float]
    base_yen: Optional[float]

    fx_jpy_per_unit: Optional[float]
    multiplier: Optional[float]
    expected_price_foreign: Optional[float]
    diff_foreign: Optional[float]


    tol_yen: Optional[float]
    tol_foreign: Optional[float]
    tol_price_digits: Optional[int]
    sku_cost_yen: Optional[float]
    buy_diff_yen: Optional[float]
    delivery_days: Optional[int]

    decision: str
    reason: str
    debug: Optional[str] = None

# -----------------------------
# ログ出力（日本語表記）
# -----------------------------
def _decision_to_ja(decision: str) -> str:
    d = (decision or "").upper()
    if d == "KEEP":
        return "保持"
    if d == "DELETE":
        return "削除"
    return decision or ""

_REASON_JA = {
    "OK": "OK",
    "MANUAL": "手動判定",
    "DELIVERY_4P": "配送日数が4日以上",
    "DELIVERY_EMPTY": "配送日数が取得できない",
    "SHIP_NOT_FOUND": "配送料が取得できない",
    "FETCH_ERROR": "自動取得失敗",
    "LISTING_PRICE_EMPTY": "B列price（外貨）が空",
    "PRICE_DIFF": "外貨差が許容差超え",
    "BUY_PRICE_MISMATCH": "仕入差NG超え",
    "SKU_COST_EMPTY": "SKU仕入が空（仕入差チェック省略）",
    "AMZ_BELOW_COST": "Amazon価格が仕入れ値より安い（優先KEEP）",
    "USED_ITEM": "中古品のため削除",
    "NO_FEATURED_OFFER": "おすすめの出品がない（BuyBox無し）",
}

def _reason_to_ja(reason_code: str) -> str:
    if reason_code is None:
        return ""
    return _REASON_JA.get(reason_code, reason_code)

def _record_to_japanese_log_row(rec: "CheckRecord") -> Dict[str, Any]:
    """CheckRecordを日本語ヘッダーの辞書に変換してCSV出力する。"""
    return {
        "判定日時": rec.timestamp,
        "国": rec.country,
        "Excel行": rec.excel_row,
        "ASIN": rec.asin,
        "SKU": rec.sku,

        "出品price(外貨)": rec.listing_price_foreign,

        "Amazon商品価格(円)": rec.amazon_item_yen,
        "Amazonページ配送料(円)": rec.amazon_page_ship_yen,
        "Amazon合計(円)": rec.amazon_total_yen,
        "出荷元Amazon": rec.shipped_by_amazon,
        "中古": rec.is_used,

        "SKU送料(円)": rec.sku_ship_yen,
        "MX小型包装物(円)": rec.mx_small_yen,
        "MX小型包装物送料(外貨)": getattr(rec, "mx_small_ship_foreign", None),
        "通関固定(円)": rec.customs_fixed_yen,
        "基準円(base)(円)": rec.base_yen,

        "為替(JPY/外貨)": rec.fx_jpy_per_unit,
        "係数(multiplier)": rec.multiplier,
        "期待価格(外貨)": rec.expected_price_foreign,
        "外貨差": rec.diff_foreign,

        "許容差(円)": rec.tol_yen,
        "許容差(外貨)": rec.tol_foreign,
        "商品価格桁数": rec.tol_price_digits,

        "SKU仕入(円)": rec.sku_cost_yen,
        "仕入差(円)": rec.buy_diff_yen,
        "配送日数": rec.delivery_days,

        "判定コード": rec.decision,
        "判定": _decision_to_ja(rec.decision),
        "理由コード": rec.reason,
        "理由": _reason_to_ja(rec.reason),

        "デバッグ": rec.debug,
    }


# 日本語ログ高速出力用：ヘッダー順と値リストを固定（DictWriterより高速）
_J_LOG_HEADERS = [
    "判定日時", "国", "Excel行", "ASIN", "SKU",
    "出品price(外貨)",
    "Amazon商品価格(円)", "Amazonページ配送料(円)", "Amazon合計(円)", "出荷元Amazon", "中古", "おすすめ出品なし",
    "SKU送料(円)", "MX小型包装物(円)", "MX小型包装物送料(外貨)", "通関固定(円)", "基準円(base)(円)",
    "為替(JPY/外貨)", "係数(multiplier)", "期待価格(外貨)", "外貨差",
    "許容差(円)", "許容差(外貨)", "商品価格桁数",
    "SKU仕入(円)", "仕入差(円)", "配送日数",
    "判定コード", "判定", "理由コード", "理由",
    "デバッグ",
]

def _record_to_japanese_log_values(rec: "CheckRecord") -> List[Any]:
    """日本語ログ用：固定順の値リストを返す（CSV書き込み高速化）。"""
    return [
        rec.timestamp, rec.country, rec.excel_row, rec.asin, rec.sku,
        rec.listing_price_foreign,
        rec.amazon_item_yen, rec.amazon_page_ship_yen, rec.amazon_total_yen, rec.shipped_by_amazon, rec.is_used,
        bool(rec.no_featured_offer) if rec.no_featured_offer is not None else False,
        rec.sku_ship_yen, rec.mx_small_yen, getattr(rec, "mx_small_ship_foreign", None), rec.customs_fixed_yen, rec.base_yen,
        rec.fx_jpy_per_unit, rec.multiplier, rec.expected_price_foreign, rec.diff_foreign,
        getattr(rec, "tol_yen", None), getattr(rec, "tol_foreign", None), getattr(rec, "tol_price_digits", None),
        rec.sku_cost_yen, rec.buy_diff_yen, rec.delivery_days,
        rec.decision, _decision_to_ja(rec.decision), rec.reason, _reason_to_ja(rec.reason),
        getattr(rec, "debug", ""),
    ]



class ListingCheckerApp(tk.Tk):

    def _load_other_settings(self):
        settings = load_settings()
        try:
            if "buy_fail_yen" in settings:
                self.var_buy_fail.set(settings["buy_fail_yen"])
            if "customs_fixed_yen" in settings:
                self.var_customs_fixed_yen.set(settings["customs_fixed_yen"])
        except Exception:
            pass


    def _load_tol_settings(self):
        settings = load_settings()
        tol = settings.get("tolerance", {})

        try:
            self.var_tol_us_4.set(tol.get("US", {}).get("4", self.var_tol_us_4.get()))
            self.var_tol_us_5.set(tol.get("US", {}).get("5", self.var_tol_us_5.get()))
            self.var_tol_us_6.set(tol.get("US", {}).get("6", self.var_tol_us_6.get()))

            self.var_tol_ca_4.set(tol.get("CA", {}).get("4", self.var_tol_ca_4.get()))
            self.var_tol_ca_5.set(tol.get("CA", {}).get("5", self.var_tol_ca_5.get()))
            self.var_tol_ca_6.set(tol.get("CA", {}).get("6", self.var_tol_ca_6.get()))

            self.var_tol_mx_4.set(tol.get("MX", {}).get("4", self.var_tol_mx_4.get()))
            self.var_tol_mx_5.set(tol.get("MX", {}).get("5", self.var_tol_mx_5.get()))
            self.var_tol_mx_6.set(tol.get("MX", {}).get("6", self.var_tol_mx_6.get()))
        except Exception:
            pass



    def _collect_tol_settings(self) -> dict:
        return {
        "US": {
            "4": self.var_tol_us_4.get(),
            "5": self.var_tol_us_5.get(),
            "6": self.var_tol_us_6.get(),
        },
        "CA": {
            "4": self.var_tol_ca_4.get(),
            "5": self.var_tol_ca_5.get(),
            "6": self.var_tol_ca_6.get(),
        },
        "MX": {
            "4": self.var_tol_mx_4.get(),
            "5": self.var_tol_mx_5.get(),
            "6": self.var_tol_mx_6.get(),
        },
    }


    def __init__(self):
        super().__init__()
        self.title("出品ファイル チェッカー（安定取得・送料込み・判定・保存）")
        self.geometry("1180x820")

        self.file_path: Optional[str] = None
        self.wb = None
        self.ws = None
        self.header_row: Optional[int] = None
        self.data_rows: List[int] = []
        self.row_to_idx: Dict[int, int] = {}

        self.current_idx: int = 0
        self.target_count: int = 10

        self.records: List[CheckRecord] = []
        self.to_delete_rows: List[int] = []

        self.buy_hard_fail_yen_default = 300.0
        self.fx_last_updated: Optional[str] = None

        self._autofill_token = 0
        self._fetch_inflight = False

        self.cur_asin: Optional[str] = None
        self.cur_item_yen: Optional[int] = None
        self.cur_page_ship_yen: Optional[int] = None
        self.cur_debug: str = ""
        self.cur_shipped_by_amazon: Optional[bool] = None
        self.cur_is_used: Optional[bool] = None
        self.cur_no_featured_offer: bool = False

        # 一括（ランダム）判定用
        self._batch_running = False
        self._batch_indices: List[int] = []
        self._batch_pos = 0
        self._batch_prev_auto_fetch = True
        self._batch_params: Dict[str, Any] = {}

        self.worker_out_q: "Queue[Dict[str, Any]]" = Queue()
        self.worker = AmazonFetchWorker(self.worker_out_q)
        self.worker.start()

        self._build_ui()
        self._load_tol_settings()
        self._load_other_settings()
        self._auto_refresh_fx()

        self.after(120, self._poll_worker_results)
        self.protocol("WM_DELETE_WINDOW", self.on_close)

    # ---------------- UI ----------------
    def _build_ui(self):
        frm_top = ttk.Frame(self)
        frm_top.pack(fill="x", padx=10, pady=8)

        ttk.Button(frm_top, text="出品ファイルを選択", command=self.choose_file).pack(side="left")

        ttk.Label(frm_top, text="  国:").pack(side="left")
        self.var_country = tk.StringVar(value="US")
        ttk.Combobox(frm_top, textvariable=self.var_country, values=["US", "CA", "MX"], width=6, state="readonly").pack(side="left")

        ttk.Label(frm_top, text="  件数:").pack(side="left", padx=(18, 0))
        self.var_target = tk.StringVar(value="10")
        ttk.Entry(frm_top, textvariable=self.var_target, width=6).pack(side="left")

        ttk.Label(frm_top, text="  仕入差NG(円):").pack(side="left", padx=(18, 0))
        self.var_buy_fail = tk.StringVar(value=str(self.buy_hard_fail_yen_default))
        ttk.Entry(frm_top, textvariable=self.var_buy_fail, width=8).pack(side="left")

        ttk.Label(frm_top, text="  通関固定(円):").pack(side="left", padx=(18, 0))
        self.var_customs_fixed_yen = tk.StringVar(value=str(CUSTOMS_FIXED_YEN_DEFAULT))
        ttk.Entry(frm_top, textvariable=self.var_customs_fixed_yen, width=8).pack(side="left")

        ttk.Label(frm_top, text="  自動取得:").pack(side="left", padx=(12, 0))
        self.var_auto_fetch = tk.BooleanVar(value=True)
        ttk.Checkbutton(frm_top, variable=self.var_auto_fetch).pack(side="left")

        ttk.Label(frm_top, text="  ブラウザ表示:").pack(side="left", padx=(12, 0))
        self.var_show_browser = tk.BooleanVar(value=(not DEFAULT_HEADLESS))
        ttk.Checkbutton(frm_top, variable=self.var_show_browser, command=self._on_toggle_browser).pack(side="left")

        ttk.Separator(self, orient="horizontal").pack(fill="x", padx=10, pady=6)

        frm_fx = ttk.Frame(self)
        frm_fx.pack(fill="x", padx=10)

        ttk.Label(frm_fx, text="為替（自動取得／30分更新）:").pack(side="left")

        self.var_usdjpy = tk.StringVar(value="")
        self.var_cadjpy = tk.StringVar(value="")
        self.var_mxnjpy = tk.StringVar(value="")

        ttk.Label(frm_fx, text=" USDJPY").pack(side="left")
        ttk.Entry(frm_fx, textvariable=self.var_usdjpy, width=10).pack(side="left", padx=(2, 10))

        ttk.Label(frm_fx, text=" CADJPY").pack(side="left")
        ttk.Entry(frm_fx, textvariable=self.var_cadjpy, width=10).pack(side="left", padx=(2, 10))

        ttk.Label(frm_fx, text=" MXNJPY").pack(side="left")
        ttk.Entry(frm_fx, textvariable=self.var_mxnjpy, width=10).pack(side="left", padx=(2, 10))

        ttk.Button(frm_fx, text="今すぐ更新", command=self.refresh_fx_now).pack(side="left", padx=10)
        self.lbl_fx_status = ttk.Label(frm_fx, text="（未取得）", foreground="gray")
        self.lbl_fx_status.pack(side="left")

        # 許容差（円） 国別設定（Amazon商品価格の桁数で自動決定）
        # 例: 商品価格が 12,345円（5桁）の場合 -> 「5桁」の許容差（円）を採用
        frm_tol = ttk.LabelFrame(self, text="許容差（円・国別）: Amazon商品価格の桁数で自動決定（最大6桁）")
        frm_tol.pack(fill="x", padx=10, pady=(6, 0))

        ttk.Label(frm_tol, text="国").grid(row=0, column=0, sticky="w", padx=(6, 0))
        ttk.Label(frm_tol, text="4桁以下（<= 9,999円）").grid(row=0, column=1, sticky="w", padx=(14, 0))
        ttk.Label(frm_tol, text="5桁（10,000〜99,999円）").grid(row=0, column=2, sticky="w", padx=(14, 0))
        ttk.Label(frm_tol, text="6桁（>= 100,000円）").grid(row=0, column=3, sticky="w", padx=(14, 0))

        def _make_tol_rule_row(row: int, code: str, d4: str = "100", d5: str = "1000", d6: str = "10000"):
            ttk.Label(frm_tol, text=code).grid(row=row, column=0, sticky="w", padx=(6, 0))
            v4 = tk.StringVar(value=d4)
            v5 = tk.StringVar(value=d5)
            v6 = tk.StringVar(value=d6)
            ttk.Entry(frm_tol, textvariable=v4, width=10).grid(row=row, column=1, sticky="w", padx=(14, 0))
            ttk.Entry(frm_tol, textvariable=v5, width=10).grid(row=row, column=2, sticky="w", padx=(14, 0))
            ttk.Entry(frm_tol, textvariable=v6, width=10).grid(row=row, column=3, sticky="w", padx=(14, 0))
            ttk.Label(frm_tol, text="円", foreground="gray").grid(row=row, column=4, sticky="w", padx=(6, 0))
            return v4, v5, v6

        self.var_tol_us_4, self.var_tol_us_5, self.var_tol_us_6 = _make_tol_rule_row(1, "US", d4="2000", d5="3000", d6="5000")
        self.var_tol_ca_4, self.var_tol_ca_5, self.var_tol_ca_6 = _make_tol_rule_row(2, "CA", d4="2200", d5="3300", d6="5400")
        self.var_tol_mx_4, self.var_tol_mx_5, self.var_tol_mx_6 = _make_tol_rule_row(3, "MX", d4="1400", d5="2100", d6="3500")

        ttk.Label(
            frm_tol,
            text="※ 判定時、Amazonの「商品価格（配送料除外）」の桁数で上記を選択し、円→外貨へ換算して許容差判定します。",
            foreground="gray"
        ).grid(row=4, column=0, columnspan=6, sticky="w", padx=(6, 0), pady=(4, 0))

        self.lbl_file = ttk.Label(self, text="（未選択）")
        self.lbl_file.pack(anchor="w", padx=12, pady=(6, 0))

        ttk.Separator(self, orient="horizontal").pack(fill="x", padx=10, pady=8)

        frm_mid = ttk.Frame(self)
        frm_mid.pack(fill="both", expand=True, padx=10)

        left = ttk.Frame(frm_mid)
        left.pack(side="left", fill="both", expand=True)

        self.lbl_progress = ttk.Label(left, text="ファイルを選択してください。", font=("Segoe UI", 12, "bold"))
        self.lbl_progress.pack(anchor="w", pady=(0, 10))

        self.txt_info = tk.Text(left, height=18)
        self.txt_info.pack(fill="both", expand=True)

        frm_inputs = ttk.Frame(left)
        frm_inputs.pack(fill="x", pady=8)

        ttk.Label(frm_inputs, text="Amazon販売価格+配送料（円）:").grid(row=0, column=0, sticky="w")
        self.var_amazon_yen = tk.StringVar()
        ttk.Entry(frm_inputs, textvariable=self.var_amazon_yen, width=18).grid(row=0, column=1, sticky="w", padx=6)

        ttk.Label(frm_inputs, text="配送日数:").grid(row=0, column=2, sticky="w", padx=(18, 0))
        self.var_delivery = tk.StringVar()
        ttk.Entry(frm_inputs, textvariable=self.var_delivery, width=10).grid(row=0, column=3, sticky="w", padx=6)

        self.lbl_calc = ttk.Label(left, text="計算: -")
        self.lbl_calc.pack(anchor="w", pady=(6, 0))

        right = ttk.Frame(frm_mid)
        right.pack(side="right", fill="y", padx=(12, 0))

        self.btn_open_amz = ttk.Button(right, text="Amazonを開く（/dp/ASIN）", command=self.open_amazon, state="disabled")
        self.btn_open_amz.pack(fill="x", pady=4)

        self.btn_autofill = ttk.Button(right, text="価格/配送を再取得（手動）", command=self.autofill_from_amazon, state="disabled")
        self.btn_autofill.pack(fill="x", pady=4)

        self.btn_eval_next = ttk.Button(right, text="判定して次へ（自動）", command=self.evaluate_and_next, state="disabled")
        self.btn_eval_next.pack(fill="x", pady=4)

        self.btn_batch_random = ttk.Button(right, text="ランダム抽出して一括判定", command=self.run_random_batch, state="disabled")
        self.btn_batch_random.pack(fill="x", pady=(10, 4))

        self.btn_keep = ttk.Button(right, text="KEEP（手動）", command=lambda: self.manual_decision("KEEP"), state="disabled")
        self.btn_keep.pack(fill="x", pady=4)

        self.btn_delete = ttk.Button(right, text="DELETE（手動）", command=lambda: self.manual_decision("DELETE"), state="disabled")
        self.btn_delete.pack(fill="x", pady=4)

        self.btn_finish = ttk.Button(right, text="終了して保存", command=self.finish_and_save, state="disabled")
        self.btn_finish.pack(fill="x", pady=(16, 4))

        self.lbl_status = ttk.Label(right, text="", foreground="gray")
        self.lbl_status.pack(fill="x", pady=(12, 0))

    def _enable_actions(self, enabled: bool):
        state = "normal" if enabled else "disabled"
        for btn in [self.btn_open_amz, self.btn_autofill, self.btn_eval_next, self.btn_batch_random,
                    self.btn_keep, self.btn_delete, self.btn_finish]:
            btn.config(state=state)

    def _set_batch_mode_ui(self, running: bool):
        if running:
            for btn in [self.btn_open_amz, self.btn_autofill, self.btn_eval_next,
                        self.btn_keep, self.btn_delete, self.btn_finish, self.btn_batch_random]:
                btn.config(state="disabled")
        else:
            self._enable_actions(True)

    def _on_toggle_browser(self):
        show = self.var_show_browser.get()
        self.worker.set_headless(headless=(not show))
        self.lbl_status.config(text="ブラウザ表示設定を変更しました。次回取得から反映されます。")

    # FX
    def refresh_fx_now(self):
        rates = fetch_fx_rates_jpy_per_unit()
        if not rates:
            messagebox.showwarning("為替取得失敗", "為替の取得に失敗しました。手動入力でも続行できます。")
            return
        self.var_usdjpy.set(f"{rates['USDJPY']:.4f}")
        self.var_cadjpy.set(f"{rates['CADJPY']:.4f}")
        self.var_mxnjpy.set(f"{rates['MXNJPY']:.4f}")
        self.fx_last_updated = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.lbl_fx_status.config(text=f"更新: {self.fx_last_updated}")

    def _auto_refresh_fx(self):
        rates = fetch_fx_rates_jpy_per_unit()
        if rates:
            self.var_usdjpy.set(f"{rates['USDJPY']:.4f}")
            self.var_cadjpy.set(f"{rates['CADJPY']:.4f}")
            self.var_mxnjpy.set(f"{rates['MXNJPY']:.4f}")
            self.fx_last_updated = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.lbl_fx_status.config(text=f"更新: {self.fx_last_updated}")
        self.after(FX_REFRESH_MS, self._auto_refresh_fx)

    def _fx_for_country(self, country: str) -> Optional[float]:
        c = country.upper()
        if c == "US":
            return try_float(self.var_usdjpy.get())
        if c == "CA":
            return try_float(self.var_cadjpy.get())
        if c == "MX":
            return try_float(self.var_mxnjpy.get())
        return None

    def _tol_rule_yen_for_country(self, country: str) -> Optional[Tuple[float, float, float]]:
        """国別の桁数ルール（4桁以下/5桁/6桁）を円で取得する。"""
        c = (country or "").strip().upper()
        if c == "US":
            v4 = try_float(getattr(self, "var_tol_us_4", tk.StringVar(value="")).get())
            v5 = try_float(getattr(self, "var_tol_us_5", tk.StringVar(value="")).get())
            v6 = try_float(getattr(self, "var_tol_us_6", tk.StringVar(value="")).get())
        elif c == "CA":
            v4 = try_float(getattr(self, "var_tol_ca_4", tk.StringVar(value="")).get())
            v5 = try_float(getattr(self, "var_tol_ca_5", tk.StringVar(value="")).get())
            v6 = try_float(getattr(self, "var_tol_ca_6", tk.StringVar(value="")).get())
        elif c == "MX":
            v4 = try_float(getattr(self, "var_tol_mx_4", tk.StringVar(value="")).get())
            v5 = try_float(getattr(self, "var_tol_mx_5", tk.StringVar(value="")).get())
            v6 = try_float(getattr(self, "var_tol_mx_6", tk.StringVar(value="")).get())
        else:
            return None
    
        if v4 is None or v5 is None or v6 is None:
            return None
        if v4 < 0 or v5 < 0 or v6 < 0:
            return None
        return float(v4), float(v5), float(v6)

    @staticmethod
    def _yen_digit_len(v: float) -> int:
        try:
            n = int(abs(float(v)))
        except Exception:
            return 0
        # 0 の場合も1桁扱い
        if n == 0:
            return 1
        return len(str(n))

    def _tol_yen_from_item_price(self, country: str, item_price_yen: Optional[float]) -> Tuple[Optional[float], Optional[int]]:
        """Amazonの商品価格（配送料除外）の桁数に応じて、許容差（円）を決める。"""
        if item_price_yen is None:
            return None, None
    
        digits = self._yen_digit_len(item_price_yen)
        # 最大6桁として扱う（7桁以上でも 6桁ルールを適用）
        if digits >= 6:
            digits_bucket = 6
        elif digits == 5:
            digits_bucket = 5
        else:
            digits_bucket = 4  # 4桁以下
    
        rule = self._tol_rule_yen_for_country(country)
        if rule is None:
            return None, digits
    
        tol4, tol5, tol6 = rule
        if digits_bucket == 6:
            return tol6, digits
        if digits_bucket == 5:
            return tol5, digits
        return tol4, digits

    # workbook
    def choose_file(self):
        path = filedialog.askopenfilename(
            title="出品ファイルを選択（xlsx / xlsm / csv）",
            filetypes=[("Excel/CSV", "*.xlsx *.xlsm *.csv"), ("All files", "*.*")]
        )
        if not path:
            return

        self.file_path = path
        self.lbl_file.config(text=f"選択中: {os.path.basename(path)}")

        t = try_float(self.var_target.get())
        self.target_count = int(t) if t is not None else 10

        try:
            self.load_workbook(path)
            self.current_idx = 0
            self.records = []
            self.to_delete_rows = []
            self.show_current_row()
            self._enable_actions(True)
        except Exception as e:
            messagebox.showerror("読み込みエラー", f"ファイルの読み込みに失敗しました。\n\n{e}")
            self._enable_actions(False)

    def load_workbook(self, path: str):
        if path.lower().endswith(".csv"):
            df = pd.read_csv(path, dtype=str, keep_default_na=False)
            tmp = os.path.join(os.path.dirname(path), f"__tmp_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            with pd.ExcelWriter(tmp, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Sheet1")
            path = tmp
            self.file_path = path
            self.lbl_file.config(text=f"選択中: {os.path.basename(path)}")

        self.wb = openpyxl.load_workbook(path)
        self.ws = self.wb.active

        self.header_row = find_header_row(self.ws)
        if not self.header_row:
            raise RuntimeError("ヘッダー行（sku/product-id）を検出できませんでした。")

        self.data_rows = build_data_row_list(self.ws, self.header_row)
        if not self.data_rows:
            raise RuntimeError("データ行を検出できませんでした。")

        self.row_to_idx = {r: i for i, r in enumerate(self.data_rows)}

    def current_excel_row(self) -> Optional[int]:
        if not self.data_rows:
            return None
        if self.current_idx < 0 or self.current_idx >= len(self.data_rows):
            return None
        return self.data_rows[self.current_idx]

    def show_current_row(self):
        r = self.current_excel_row()
        if r is None:
            self.lbl_progress.config(text="対象行がありません（終了して保存してください）。")
            self.txt_info.delete("1.0", tk.END)
            self.lbl_calc.config(text="計算: -")
            return

        country = self.var_country.get().strip().upper()
        sku = clean_text(self.ws.cell(r, COL_SKU).value)
        asin = clean_text(self.ws.cell(r, COL_ASIN).value).replace(" ", "")

        self.cur_asin = asin
        self.cur_item_yen = None
        self.cur_page_ship_yen = None
        self.cur_debug = ""
        self.cur_shipped_by_amazon = None
        self.cur_is_used = None

        price_foreign = try_float(self.ws.cell(r, COL_PRICE).value)
        qty = clean_text(self.ws.cell(r, COL_QTY).value)

        sku_cost = parse_sku_cost_yen(sku)
        ship_yen, mx_small_yen = parse_sku_shipping_yen(sku, country)

        self.lbl_progress.config(text=f"国:{country}  進捗:{len(self.records)}/{self.target_count}  （Excel行:{r}）")

        self.txt_info.delete("1.0", tk.END)
        self.txt_info.insert(tk.END, f"SKU:\n  {sku}\n\n")
        self.txt_info.insert(tk.END, f"SKU仕入（円）:\n  {sku_cost}\n\n")
        self.txt_info.insert(tk.END, f"SKU送料（円）:\n  {ship_yen}\n")
        self.txt_info.insert(tk.END, f"MX小型包装物値（円）:\n  {mx_small_yen}\n\n")
        self.txt_info.insert(tk.END, f"ASIN(product-id):\n  {asin}\n\n")
        self.txt_info.insert(tk.END, f"B列 price（外貨）:\n  {price_foreign}\n\n")
        self.txt_info.insert(tk.END, f"quantity:\n  {qty}\n\n")

        self.var_amazon_yen.set("")
        self.var_delivery.set("")
        self.lbl_calc.config(text="計算: -")
        self.lbl_status.config(text="")

        if self.var_auto_fetch.get() and (not self._batch_running):
            self._schedule_autofill()

    def _schedule_autofill(self):
        # 既存の取得を無効化するためトークンだけ進める
        self._autofill_token += 1
        token = self._autofill_token
        self.after(200, lambda: self._autofill_if_still_current(token))

    def _autofill_if_still_current(self, token: int):
        if token != self._autofill_token:
            return
        if self._fetch_inflight:
            return
        self.autofill_from_amazon(token=token)

    def _autofill_timeout(self, token: int):
        if token != self._autofill_token:
            return
        if not self._fetch_inflight:
            return
        self._autofill_token += 1
        self._fetch_inflight = False

        # バッチ実行中は手動ボタンを再有効化しない（誤操作防止）
        if not self._batch_running:
            self.btn_autofill.config(state="normal")
        self.lbl_status.config(text="自動取得がタイムアウトしました。手入力または『再取得（手動）』をお試しください。")
        if self._batch_running:
            # タイムアウトはDELETEとして継続
            self.after(10, lambda: self._batch_finalize_one(fetch_ok=False, debug="timeout"))

    def autofill_from_amazon(self, token: Optional[int] = None):
        r = self.current_excel_row()
        if r is None:
            return
        asin = clean_text(self.ws.cell(r, COL_ASIN).value).replace(" ", "")
        if not ASIN_RE.match(asin):
            self.lbl_status.config(text=f"ASIN形式エラー: {asin}")
            if self._batch_running:
                self.after(10, lambda: self._batch_finalize_one(fetch_ok=False, debug="asin_format"))
            return

        if token is None:
            self._autofill_token += 1
            token = self._autofill_token
        else:
            self._autofill_token = token

        # SKU仕入れ値（出品一覧から候補選定に使用）

        sku = clean_text(self.ws.cell(r, COL_SKU).value)

        sku_cost_yen = parse_sku_cost_yen(sku)


        self._fetch_inflight = True

        self.after(AMZ_HARD_TIMEOUT_MS, lambda: self._autofill_timeout(token))
        self.btn_autofill.config(state="disabled")
        self.lbl_status.config(text="自動取得中...")

        self.worker.fetch(asin=asin, token=token, sku_cost_yen=sku_cost_yen)

    def _poll_worker_results(self):
        try:
            while True:
                res = self.worker_out_q.get_nowait()
                self._apply_fetch_result(res)
        except Empty:
            pass
        except Exception:
            pass
        self.after(120, self._poll_worker_results)

    def _apply_fetch_result(self, res: Dict[str, Any]):
        token = int(res.get("token", -1))
        if token != self._autofill_token:
            return

        self._fetch_inflight = False
        if not self._batch_running:
            self.btn_autofill.config(state="normal")

        asin = res.get("asin", "")
        ok = res.get("ok", False)
        item = res.get("item", None)
        ship = res.get("ship", None)
        days = res.get("days", None)
        debug = res.get("debug", "")
        shipped_by_amazon = res.get("shipped_by_amazon", None)
        is_used = res.get("is_used", None)
        no_featured_offer = bool(res.get('no_featured_offer', False))

        self.cur_asin = asin
        self.cur_item_yen = item
        self.cur_page_ship_yen = ship
        self.cur_debug = debug
        self.cur_shipped_by_amazon = shipped_by_amazon
        self.cur_is_used = is_used
        self.cur_no_featured_offer = no_featured_offer

        # 送料が取れない場合は total を確定できないため、合計欄は空にする（評価はDELETE）
        total = None
        if item is not None and ship is not None:
            total = item + ship

        if total is not None:
            self.var_amazon_yen.set(str(total))
        else:
            self.var_amazon_yen.set("")

        if days is not None:
            self.var_delivery.set(str(days))
        else:
            self.var_delivery.set("")

        if ship is None:
            ship_msg = "配送料=未取得（この行はDELETE判定）"
        else:
            ship_msg = f"配送料={ship}"

        if not ok and total is None and days is None:
            self.lbl_status.config(text=f"自動取得失敗: {debug}")
        else:
            self.lbl_status.config(text=f"自動取得: 商品={item}, {ship_msg}, 合計={total}, 配送日数={days}, 出荷元Amazon={shipped_by_amazon}, 中古={is_used}, おすすめ出品なし={no_featured_offer} / {debug}")

        if self._batch_running:
            self.after(10, lambda: self._batch_finalize_one(fetch_ok=bool(ok), debug=debug))

    def open_amazon(self):
        r = self.current_excel_row()
        if r is None:
            return
        asin = clean_text(self.ws.cell(r, COL_ASIN).value).replace(" ", "")
        webbrowser.open_new_tab(f"https://www.amazon.co.jp/dp/{asin}")

    # ---- 判定ロジック（共通） ----
    def _compute_record_for_row(
        self,
        r: int,
        country: str,
        fx: float,
        buy_fail: float,
        customs_fixed_yen: float,
        amazon_total_yen_input: Optional[float],
        delivery_days_input: Optional[int],
        amazon_item_yen: Optional[int],
        amazon_page_ship_yen: Optional[int],
        shipped_by_amazon: Optional[bool] = None,
        is_used: Optional[bool] = None,
        no_featured_offer: bool = False,
    ) -> CheckRecord:
        sku = clean_text(self.ws.cell(r, COL_SKU).value)
        asin = clean_text(self.ws.cell(r, COL_ASIN).value).replace(" ", "")
        listing_price = try_float(self.ws.cell(r, COL_PRICE).value)
    
        ship_yen, mx_small_yen = parse_sku_shipping_yen(sku, country)
        # mx_small_yen は「小型包装物の値（円）」であり、base_yen には加算しない（MXのみ別式で外貨側へ加算）
        sku_ship_total_yen = ship_yen
        sku_cost = parse_sku_cost_yen(sku)
    
        multiplier = COUNTRY_MULTIPLIER.get(country, 1.0)
        ts = dt.datetime.now().isoformat(timespec="seconds")
    
        decision = "DELETE"
        reason = ""
        buy_diff = None
        base_yen = None
        expected = None
        diff_foreign = None
        mx_small_ship_foreign = None
    
        amazon_total_yen_used = amazon_total_yen_input
    
        # --- 許容差（桁数ルール） ---
        # Amazonの商品価格（配送料除外）を優先し、取れなければ total - ship で推定。
        tol_item_yen = None
        if amazon_item_yen is not None:
            tol_item_yen = float(amazon_item_yen)
        elif amazon_total_yen_input is not None and amazon_page_ship_yen is not None:
            tol_item_yen = float(amazon_total_yen_input) - float(amazon_page_ship_yen)
        elif amazon_total_yen_input is not None:
            tol_item_yen = float(amazon_total_yen_input)
    
        tol_yen, tol_digits = self._tol_yen_from_item_price(country, tol_item_yen)
        tol_foreign = (tol_yen / fx) * multiplier if (tol_yen is not None and fx > 0) else None
    
        delivery_override = bool(shipped_by_amazon)
        delivery_days_int = int(delivery_days_input) if delivery_days_input is not None else None
        delivery_override_triggered = bool(delivery_override and (delivery_days_int is None or delivery_days_int >= 4))
        # --- 参考計算（ログ用） ---
        # 判定が配送日数などで早期DELETEになっても、可能な範囲で base/expected/MX小型包装物送料 を計算してログに残す。
        amazon_total_yen_calc = amazon_total_yen_input
        if amazon_total_yen_calc is None:
            if amazon_item_yen is not None and amazon_page_ship_yen is not None:
                amazon_total_yen_calc = float(amazon_item_yen + amazon_page_ship_yen)
        if amazon_total_yen_calc is not None:
            amazon_total_yen_used = amazon_total_yen_calc

        buy_base_yen_calc = None
        if amazon_item_yen is not None:
            buy_base_yen_calc = float(amazon_item_yen)
        elif amazon_total_yen_calc is not None and amazon_page_ship_yen is not None:
            buy_base_yen_calc = float(amazon_total_yen_calc) - float(amazon_page_ship_yen)

        if amazon_total_yen_calc is not None and fx > 0:
            base_yen = float(amazon_total_yen_calc) + float(sku_ship_total_yen) + float(customs_fixed_yen)
            expected_base = (base_yen / fx) * multiplier
            if country.upper() == "MX" and mx_small_yen and float(mx_small_yen) > 0 and buy_base_yen_calc is not None:
                mx_small_ship_foreign = round(((buy_base_yen_calc + float(mx_small_yen)) * 1.35) / fx, 4)
                expected = round(expected_base + mx_small_ship_foreign, 2)
            else:
                expected = round(expected_base, 2)
            if listing_price is not None:
                diff_foreign = round(abs(listing_price - expected), 4)


        # (0) 中古品は即DELETE（仕様追加）
        if is_used is True:
            decision = "DELETE"
            reason = "USED_ITEM"
        # (1) 送料未取得は即DELETE（仕様）
        # (0.5) おすすめの出品（BuyBox）が無い場合はDELETE（仕様変更）
        elif bool(no_featured_offer):
            decision = "DELETE"
            reason = "NO_FEATURED_OFFER"
        elif amazon_page_ship_yen is None:
            decision = "DELETE"
            reason = "SHIP_NOT_FOUND"
        # (2) 配送日数（出荷元Amazonなら免除）
        elif (not delivery_override) and (delivery_days_int is None):
            decision = "DELETE"
            reason = "DELIVERY_EMPTY"
        elif (not delivery_override) and (delivery_days_int >= 4):
            decision = "DELETE"
            reason = "DELIVERY_4P"
        # (3) 許容差ルールが不正ならDELETE（通常はUI側で弾くが保険）
        elif tol_yen is None or tol_foreign is None:
            decision = "DELETE"
            reason = "TOL_RULE_INVALID"
        else:
            # amazon_total_yen を確定（入力優先、なければ item+ship）
            amazon_total_yen = amazon_total_yen_input
            if amazon_total_yen is None:
                if amazon_item_yen is not None and amazon_page_ship_yen is not None:
                    amazon_total_yen = float(amazon_item_yen + amazon_page_ship_yen)
                else:
                    amazon_total_yen = None
    
            amazon_total_yen_used = amazon_total_yen
    
            if amazon_total_yen is None:
                decision = "DELETE"
                reason = "AMAZON_TOTAL_EMPTY"
            else:
                # 仕入差比較は「商品本体価格（配送料除外）」が取れていればそれで比較
                if amazon_item_yen is not None:
                    buy_base_yen = float(amazon_item_yen)
                else:
                    buy_base_yen = float(amazon_total_yen - float(amazon_page_ship_yen))
    
                cheaper_override = False
                if sku_cost is None:
                    buy_ok = True
                    buy_reason = "SKU_COST_EMPTY"
                else:
                    buy_diff = round(abs(buy_base_yen - sku_cost), 2)
                    # 仕様変更: Amazon商品価格がSKU仕入れ値より安い場合はKEEP扱い
                    if buy_base_yen < sku_cost:
                        cheaper_override = True
                        buy_ok = True
                        buy_reason = "AMZ_BELOW_COST"
                    else:
                        buy_ok = buy_diff <= buy_fail
                        buy_reason = "OK" if buy_ok else "BUY_PRICE_MISMATCH"
    
                base_yen = float(amazon_total_yen) + float(sku_ship_total_yen) + float(customs_fixed_yen)

                # MXのみ：housou/kogata の場合は小型包装物送料（外貨）を追加
                # 小型包装物送料(外貨) = (Amazon商品価格（配送料除外） + SKU小型包装物値) × 1.35 ÷ MXNJPY
                expected_base = (base_yen / fx) * multiplier
                if country.upper() == "MX" and mx_small_yen and float(mx_small_yen) > 0:
                    mx_small_ship_foreign = round(((buy_base_yen + float(mx_small_yen)) * 1.35) / fx, 4)
                    expected = round(expected_base + mx_small_ship_foreign, 2)
                else:
                    expected = round(expected_base, 2)
    
                if listing_price is None:
                    decision = "DELETE"
                    reason = "LISTING_PRICE_EMPTY"
                else:
                    diff_foreign = round(abs(listing_price - expected), 4)
                    # 仕様変更: Amazon商品価格がSKU仕入れ値より安い場合は、外貨差/仕入差を問わずKEEP
                    if 'cheaper_override' in locals() and cheaper_override:
                        decision = "KEEP"
                        reason = "AMZ_BELOW_COST"
                    elif not buy_ok:
                        decision = "DELETE"
                        reason = buy_reason
                    else:
                        decision = "KEEP" if diff_foreign <= tol_foreign else "DELETE"
                        reason = "OK" if diff_foreign <= tol_foreign else "PRICE_DIFF"
    
        rec = CheckRecord(
            timestamp=ts,
            country=country,
            excel_row=r,
            asin=asin,
            sku=sku,
            listing_price_foreign=listing_price,
            amazon_item_yen=float(amazon_item_yen) if amazon_item_yen is not None else None,
            amazon_page_ship_yen=float(amazon_page_ship_yen) if amazon_page_ship_yen is not None else None,
            amazon_total_yen=amazon_total_yen_used,
            shipped_by_amazon=shipped_by_amazon,
            is_used=is_used,
            no_featured_offer=bool(no_featured_offer),
            sku_ship_yen=ship_yen,
            mx_small_yen=mx_small_yen,
            mx_small_ship_foreign=mx_small_ship_foreign,
            customs_fixed_yen=customs_fixed_yen,
            base_yen=base_yen,
            fx_jpy_per_unit=fx,
            multiplier=multiplier,
            expected_price_foreign=expected,
            diff_foreign=diff_foreign,
            tol_yen=float(tol_yen) if tol_yen is not None else None,
            tol_foreign=float(tol_foreign) if tol_foreign is not None else None,
            tol_price_digits=int(tol_digits) if tol_digits is not None else None,
            sku_cost_yen=sku_cost,
            buy_diff_yen=buy_diff,
            delivery_days=int(delivery_days_input) if delivery_days_input is not None else None,
            decision=decision,
            reason=reason,
            debug=(self.cur_debug if self.cur_asin == asin else None),
        )
        return rec

    def _evaluate_current_row(self, show_messages: bool, advance_sequential: bool):
        r = self.current_excel_row()
        if r is None:
            return
    
        country = self.var_country.get().strip().upper()
        fx = self._fx_for_country(country)
        tol_rule = self._tol_rule_yen_for_country(country)
        buy_fail = try_float(self.var_buy_fail.get())
        customs_fixed_yen = try_float(self.var_customs_fixed_yen.get())
    
        if fx is None or fx <= 0:
            if show_messages:
                messagebox.showerror("入力エラー", "為替が未取得/不正です。USDJPY/CADJPY/MXNJPYを確認してください。")
            else:
                self.lbl_status.config(text="為替が未取得/不正のため一括判定を中断しました。")
                self._batch_abort()
            return
        if tol_rule is None:
            if show_messages:
                messagebox.showerror("入力エラー", "許容差（円）の桁数ルールが不正です（国別に 4桁/5桁/6桁 を入力してください）。")
            else:
                self.lbl_status.config(text="許容差（円）の桁数ルールが不正のため一括判定を中断しました。")
                self._batch_abort()
            return
        if buy_fail is None or buy_fail < 0:
            if show_messages:
                messagebox.showerror("入力エラー", "仕入差NG(円)が正しくありません。")
            else:
                self.lbl_status.config(text="仕入差NG(円)が不正のため一括判定を中断しました。")
                self._batch_abort()
            return
        if customs_fixed_yen is None or customs_fixed_yen < 0:
            if show_messages:
                messagebox.showerror("入力エラー", "通関固定(円)が正しくありません。")
            else:
                self.lbl_status.config(text="通関固定(円)が不正のため一括判定を中断しました。")
                self._batch_abort()
            return
    
        # 入力（Amazon合計・配送日数）
        amazon_total_yen_input = try_float(self.var_amazon_yen.get())
        delivery_days_f = try_float(self.var_delivery.get())
        delivery_days_input = int(delivery_days_f) if delivery_days_f is not None else None
    
        asin = clean_text(self.ws.cell(r, COL_ASIN).value).replace(" ", "")
        amazon_item_yen = self.cur_item_yen if self.cur_asin == asin else None
        amazon_page_ship_yen = self.cur_page_ship_yen if self.cur_asin == asin else None
        shipped_by_amazon = self.cur_shipped_by_amazon if self.cur_asin == asin else None
        is_used = self.cur_is_used if self.cur_asin == asin else None
    
        rec = self._compute_record_for_row(
            r=r,
            country=country,
            fx=fx,
            buy_fail=buy_fail,
            customs_fixed_yen=customs_fixed_yen,
            amazon_total_yen_input=amazon_total_yen_input,
            delivery_days_input=delivery_days_input,
            amazon_item_yen=amazon_item_yen,
            amazon_page_ship_yen=amazon_page_ship_yen,
            shipped_by_amazon=shipped_by_amazon,
            is_used=is_used,
            no_featured_offer=self.cur_no_featured_offer if self.cur_asin == asin else False,
        )
    
        multiplier = COUNTRY_MULTIPLIER.get(country, 1.0)
    
        tol_yen_disp = "-" if rec.tol_yen is None else f"{int(rec.tol_yen):,}円"
        tol_foreign_disp = "-" if rec.tol_foreign is None else f"{rec.tol_foreign:.4f}"
        digits_disp = "-" if rec.tol_price_digits is None else str(rec.tol_price_digits)

        mx_small_ship_disp = ""
        if country == "MX" and getattr(rec, "mx_small_ship_foreign", None):
            mx_small_ship_disp = f" | MX小型包装物送料(外貨)={rec.mx_small_ship_foreign}"
    
        self.lbl_calc.config(
            text=f"Amazon合計={amazon_total_yen_input}（商品={amazon_item_yen}, 配送料={amazon_page_ship_yen}）"
                 f" | 通関={customs_fixed_yen}"
                 f" | 許容差(digits={digits_disp})={tol_yen_disp} -> {tol_foreign_disp}"
                 f" | FX={fx:.4f} | 係数={multiplier}"
                 f" => base={rec.base_yen} | Expected={rec.expected_price_foreign}{mx_small_ship_disp}"
                 f" | 外貨差={rec.diff_foreign}"
                 f" | 仕入差={rec.buy_diff_yen}（NG>{buy_fail}）"
        )
        self.lbl_status.config(text=f"判定: {rec.decision}（{rec.reason}）")
    
        self.records.append(rec)
        if rec.decision == "DELETE":
            self.to_delete_rows.append(r)
    
        if advance_sequential:
            self.current_idx += 1
            self.show_current_row()

    # 判定（UIボタン）
    def evaluate_and_next(self):
        if self._batch_running:
            return

        if len(self.records) >= self.target_count:
            messagebox.showinfo("完了", "指定件数に到達しました。終了して保存してください。")
            return

        # 送料が未取得でもエラーにせずDELETEにする（仕様変更に合わせる）
        self._evaluate_current_row(show_messages=True, advance_sequential=True)

        if len(self.records) >= self.target_count:
            messagebox.showinfo("完了", "指定件数に到達しました。終了して保存してください。")
            self.btn_eval_next.config(state="disabled")

    def manual_decision(self, decision: str):
        if self._batch_running:
            return

        if len(self.records) >= self.target_count:
            messagebox.showinfo("完了", "指定件数に到達しました。終了して保存してください。")
            return

        r = self.current_excel_row()
        if r is None:
            return

        country = self.var_country.get().strip().upper()
        fx = self._fx_for_country(country)

        sku = clean_text(self.ws.cell(r, COL_SKU).value)
        asin = clean_text(self.ws.cell(r, COL_ASIN).value).replace(" ", "")
        listing_price = try_float(self.ws.cell(r, COL_PRICE).value)

        amazon_total_yen = try_float(self.var_amazon_yen.get())
        delivery_days_f = try_float(self.var_delivery.get())
        delivery_days = int(delivery_days_f) if delivery_days_f is not None else None

        ship_yen, mx_small_yen = parse_sku_shipping_yen(sku, country)
        customs_fixed_yen = try_float(self.var_customs_fixed_yen.get())
        sku_cost = parse_sku_cost_yen(sku)

        ts = dt.datetime.now().isoformat(timespec="seconds")

        rec = CheckRecord(
            timestamp=ts,
            country=country,
            excel_row=r,
            asin=asin,
            sku=sku,
            listing_price_foreign=listing_price,
            amazon_item_yen=float(self.cur_item_yen) if (self.cur_asin == asin and self.cur_item_yen is not None) else None,
            amazon_page_ship_yen=float(self.cur_page_ship_yen) if (self.cur_asin == asin and self.cur_page_ship_yen is not None) else None,
            amazon_total_yen=amazon_total_yen,
            shipped_by_amazon=(self.cur_shipped_by_amazon if self.cur_asin == asin else None),
            is_used=(self.cur_is_used if self.cur_asin == asin else None),
            sku_ship_yen=ship_yen,
            mx_small_yen=mx_small_yen,
            mx_small_ship_foreign=None,
            customs_fixed_yen=customs_fixed_yen,
            base_yen=None,
            fx_jpy_per_unit=fx,
            multiplier=COUNTRY_MULTIPLIER.get(country, None),
            expected_price_foreign=None,
            diff_foreign=None,
            tol_yen=None,
            tol_foreign=None,
            tol_price_digits=None,
            sku_cost_yen=sku_cost,
            buy_diff_yen=None,
            delivery_days=delivery_days,
            decision=decision,
            reason="MANUAL"
        )

        self.records.append(rec)
        if decision == "DELETE":
            self.to_delete_rows.append(r)

        self.current_idx += 1
        self.show_current_row()

    # ---- ランダム一括判定 ----

    def _validate_params_for_batch(self) -> Optional[Dict[str, Any]]:
        country = self.var_country.get().strip().upper()
        fx = self._fx_for_country(country)
        tol_rule = self._tol_rule_yen_for_country(country)
        buy_fail = try_float(self.var_buy_fail.get())
        customs_fixed_yen = try_float(self.var_customs_fixed_yen.get())
        t = try_float(self.var_target.get())
        target = int(t) if t is not None else 10
    
        if fx is None or fx <= 0:
            messagebox.showerror("入力エラー", "為替が未取得/不正です。USDJPY/CADJPY/MXNJPYを確認してください。")
            return None
        if tol_rule is None:
            messagebox.showerror("入力エラー", "許容差（円）の桁数ルールが不正です（国別に 4桁/5桁/6桁 を入力してください）。")
            return None
        if buy_fail is None or buy_fail < 0:
            messagebox.showerror("入力エラー", "仕入差NG(円)が正しくありません。")
            return None
        if customs_fixed_yen is None or customs_fixed_yen < 0:
            messagebox.showerror("入力エラー", "通関固定(円)が正しくありません。")
            return None
        if target <= 0:
            messagebox.showerror("入力エラー", "件数が正しくありません。")
            return None
    
        return {
            "country": country,
            "fx": fx,
            "buy_fail": buy_fail,
            "customs_fixed_yen": customs_fixed_yen,
            "target": target,
        }

    def run_random_batch(self):
        if self._batch_running:
            return
        if not self.ws or not self.data_rows:
            return

        params = self._validate_params_for_batch()
        if not params:
            return

        # 対象行（ASINが妥当な行）の中からランダム抽出
        eligible = []
        for i, rr in enumerate(self.data_rows):
            asin = clean_text(self.ws.cell(rr, COL_ASIN).value).replace(" ", "")
            if ASIN_RE.match(asin):
                eligible.append(i)

        if not eligible:
            messagebox.showwarning("対象なし", "ASINが有効な行を検出できませんでした。")
            return

        k = min(int(params["target"]), len(eligible))
        self.target_count = k
        self._batch_indices = random.sample(eligible, k)
        self._batch_pos = 0
        self._batch_running = True
        self._batch_prev_auto_fetch = bool(self.var_auto_fetch.get())
        self._batch_params = params
        # 一括開始前に、未処理の自動取得予約／タイムアウトを無効化（トークンでキャンセル）
        # これにより、直前の自動取得が残っていてもバッチ処理に干渉しません。
        self._autofill_token += 1
        self._fetch_inflight = False


        # 一括は自動取得OFFで制御（自動取得は自前で回す）
        self.var_auto_fetch.set(False)

        # 記録は一旦リセット（「ランダム10件だけ判定」用途）
        self.records = []
        self.to_delete_rows = []

        self.lbl_status.config(text=f"ランダム抽出 {k} 件の一括判定を開始します...")
        self._set_batch_mode_ui(True)

        self.after(50, self._batch_go_next)

    def _batch_abort(self):
        if not self._batch_running:
            return
        self._batch_running = False
        self.var_auto_fetch.set(self._batch_prev_auto_fetch)
        self._set_batch_mode_ui(False)
        self.btn_finish.config(state="normal" if self.file_path else "disabled")
        self.lbl_status.config(text="一括判定を中断しました。")

    def _batch_go_next(self):
        if not self._batch_running:
            return

        if self._batch_pos >= len(self._batch_indices):
            self._batch_finish()
            return

        idx = self._batch_indices[self._batch_pos]
        self.current_idx = idx
        self.show_current_row()

        # 直前の取得がまだ終わっている場合は待機
        if self._fetch_inflight:
            self.after(120, self._batch_go_next)
            return

        # 取得開始（tokenは都度新規）
        self.autofill_from_amazon(token=None)

    def _batch_finalize_one(self, fetch_ok: bool, debug: str):
        """
        取得結果が反映されたタイミングで、現在行を評価し次へ進む。
        """
        if not self._batch_running:
            return

        # 取得そのものに失敗した場合は「DELETE」ではなく「SKIP」としてログに残し、次へ進む。
        # （コード不整合/一時的なPlaywright不調で一括が全DELETEになる事故を防ぐ）
        if not fetch_ok:
            try:
                r = self.current_excel_row()
                country = self.var_country.get().strip().upper()
                sku = clean_text(self.ws.cell(r, COL_SKU).value) if (r is not None and self.ws is not None) else ""
                asin = clean_text(self.ws.cell(r, COL_ASIN).value).replace(" ", "") if (r is not None and self.ws is not None) else ""
                listing_price = try_float(self.ws.cell(r, COL_PRICE).value) if (r is not None and self.ws is not None) else None
                ts = dt.datetime.now().isoformat(timespec="seconds")

                rec = CheckRecord(
                    timestamp=ts,
                    country=country,
                    excel_row=int(r) if r is not None else -1,
                    asin=asin,
                    sku=sku,
                    listing_price_foreign=listing_price,

                    amazon_item_yen=None,
                    amazon_page_ship_yen=None,
                    amazon_total_yen=None,
                    shipped_by_amazon=None,
                    is_used=None,

                    sku_ship_yen=None,
                    mx_small_yen=None,
                    mx_small_ship_foreign=None,
                    customs_fixed_yen=None,
                    base_yen=None,

                    fx_jpy_per_unit=self._fx_for_country(country),
                    multiplier=COUNTRY_MULTIPLIER.get(country, None),
                    expected_price_foreign=None,
                    diff_foreign=None,

                    tol_yen=None,
                    tol_foreign=None,
                    tol_price_digits=None,
                    sku_cost_yen=parse_sku_cost_yen(sku),
                    buy_diff_yen=None,
                    delivery_days=None,

                    decision="SKIP",
                    reason="FETCH_ERROR",
                    debug=debug,
                )
                self.records.append(rec)
            except Exception:
                pass

            self.lbl_status.config(text=f"自動取得失敗のためスキップ: {debug}")
            self._batch_pos += 1
            self.after(120, self._batch_go_next)
            return

        # 送料未取得は仕様上 DELETE

        try:
            self._evaluate_current_row(show_messages=False, advance_sequential=False)
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            try:
                with open('batch_error.log', 'a', encoding='utf-8') as f:
                    f.write(tb + '\n')
            except Exception:
                pass
            # バッチを解除してUIを戻す
            self.lbl_status.config(text=f"一括判定中にエラーが発生しました: {e}")
            self._batch_abort()
            try:
                messagebox.showerror('一括判定エラー', f"一括判定中にエラーが発生したため中断しました。\n\n{e}\n\n詳細は batch_error.log を確認してください。")
            except Exception:
                pass
            return

        self._batch_pos += 1
        self.after(120, self._batch_go_next)

    def _batch_finish(self):
        self._batch_running = False
        self.var_auto_fetch.set(self._batch_prev_auto_fetch)

        kept = sum(1 for r in self.records if r.decision == "KEEP")
        deleted = sum(1 for r in self.records if r.decision == "DELETE")

        self._set_batch_mode_ui(False)
        self.btn_finish.config(state="normal")
        self.lbl_status.config(text=f"一括判定完了: {len(self.records)}件（KEEP={kept}, DELETE={deleted}）。必要なら『終了して保存』してください。")

    # 保存
    def finish_and_save(self):
        if not self.file_path:
            return

        if not self.records:
            if not messagebox.askyesno("確認", "記録がありません。終了しますか？"):
                return

        base_dir = os.path.dirname(self.file_path)
        base_name = os.path.splitext(os.path.basename(self.file_path))[0]
        stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        country = self.var_country.get().strip().upper()

        out_xlsx = os.path.join(base_dir, f"{base_name}_checked_{country}_{stamp}.xlsx")
        out_log = os.path.join(base_dir, f"{base_name}_log_{country}_{stamp}.csv")

        self.lbl_status.config(text="保存中...（Excelとログを書き出しています）")
        try:
            self.update_idletasks()
        except Exception:
            pass


        shutil.copy2(self.file_path, out_xlsx)
        wb2 = openpyxl.load_workbook(out_xlsx)
        ws2 = wb2.active

        rows_del = sorted(set(self.to_delete_rows))
        if rows_del:
            # 連続行はまとめて削除（openpyxlのdelete_rows呼び出し回数を減らして高速化）
            groups: List[Tuple[int, int]] = []
            s = rows_del[0]
            prev = rows_del[0]
            for rdel in rows_del[1:]:
                if rdel == prev + 1:
                    prev = rdel
                    continue
                groups.append((s, prev - s + 1))
                s = prev = rdel
            groups.append((s, prev - s + 1))
            for s, n in reversed(groups):
                ws2.delete_rows(s, n)

        wb2.save(out_xlsx)
        wb2.close()

        
        if self.records:
            # ログは日本語ヘッダー＋日本語判定/理由で出力（コードも併記）
            # DictWriter＋全件dict生成は遅くなりやすいので、固定順のcsv.writerで逐次書き出し（高速化）
            with open(out_log, "w", newline="", encoding="utf-8-sig", buffering=1024 * 1024) as f:
                w = csv.writer(f)
                w.writerow(_J_LOG_HEADERS)
                for rec in self.records:
                    w.writerow(_record_to_japanese_log_values(rec))
        kept = sum(1 for r in self.records if r.decision == "KEEP")
        deleted = sum(1 for r in self.records if r.decision == "DELETE")

        messagebox.showinfo(
            "保存完了",
            f"保存しました。\n\n"
            f"削除反映済み: {out_xlsx}\n"
            f"ログ: {out_log}\n\n"
            f"KEEP={kept}, DELETE={deleted}"
        )

    def on_close(self):
        try:
            save_settings({
                "tolerance": self._collect_tol_settings(),
                "buy_fail_yen": self.var_buy_fail.get(),
                "customs_fixed_yen": self.var_customs_fixed_yen.get(),
            })
            self.worker.stop()
        except Exception:
            pass
        finally:
            self.destroy()


if __name__ == "__main__":
    import traceback
    try:
        app = ListingCheckerApp()
        app.mainloop()
    except SystemExit:
        raise
    except Exception:
        tb = traceback.format_exc()
        # 起動直後に落ちてコンソールが閉じても原因が分かるようにログ出力
        try:
            with open("startup_error.log", "w", encoding="utf-8") as f:
                f.write(tb)
        except Exception:
            pass
        try:
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror(
                "起動エラー",
                "起動に失敗しました。\n\n同じフォルダに startup_error.log を出力しました。\n\n" + tb[-1800:]
            )
            root.destroy()
        except Exception:
            print(tb)